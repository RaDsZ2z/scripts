#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""Local browser UI for the two Nano Banana image generators."""

from __future__ import annotations

import argparse
import base64
import datetime as dt
import hashlib
import hmac
import io
import json
import mimetypes
import os
import re
import secrets
import threading
import time
import uuid
import webbrowser
from concurrent.futures import ThreadPoolExecutor, as_completed
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from http.cookies import SimpleCookie
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urlparse

import requests

from api_key_config import get_api_key


ROOT = Path(__file__).resolve().parent
WEB_ROOT = ROOT / "web"
MAX_BODY_BYTES = 300 * 1024 * 1024
GLOBAL_API_CONCURRENCY = 5
MAX_WORKERS_PER_JOB = 2
JOB_RETENTION_SECONDS = 2 * 60 * 60
JOB_CLEANUP_INTERVAL_SECONDS = 5 * 60
SESSION_COOKIE_NAME = "banana_session"
SESSION_TTL_SECONDS = 12 * 60 * 60

PROVIDERS = {
    "ikun": {
        "label": "Ikun Banana",
        "script": ROOT / "ikun_banana.py",
        "env": "IKUN_BANANA_API_KEY",
        "api_url": "https://api.ikuncode.cc/v1/chat/completions",
        "model": "gemini-3.1-flash-image-preview",
        "api_style": "chat",
    },
    "micu": {
        "label": "Micu Gpt Image2",
        "script": ROOT / "micu_gpt_image_2.py",
        "env": "MICU_API_KEY",
        "api_url": "https://www.micuapi.ai",
        "model": "gpt-image-2",
        "api_style": "images",
    },
    "banana": {
        "label": "Laozhang Banana",
        "script": ROOT / "laozhang_banana.py",
        "env": "BANANA_API_KEY",
        "api_url": "https://api2.laozhang.ai/v1/chat/completions",
        "model": "gemini-3.1-flash-image",
        "api_style": "chat",
    },
}

JOBS: dict[str, dict[str, Any]] = {}
JOBS_LOCK = threading.Lock()
API_REQUEST_SLOTS = threading.BoundedSemaphore(GLOBAL_API_CONCURRENCY)


def now_iso() -> str:
    return dt.datetime.now().isoformat(timespec="seconds")


def auth_enabled() -> bool:
    return bool(os.environ.get("BANANA_WEB_PASSWORD_HASH") and os.environ.get("BANANA_WEB_SESSION_SECRET"))


def decode_urlsafe(value: str) -> bytes:
    return base64.urlsafe_b64decode(value + "=" * (-len(value) % 4))


def password_matches(password: str) -> bool:
    try:
        algorithm, iterations_text, salt_text, expected_text = os.environ["BANANA_WEB_PASSWORD_HASH"].split("$", 3)
        if algorithm != "pbkdf2_sha256":
            return False
        actual = hashlib.pbkdf2_hmac(
            "sha256",
            password.encode("utf-8"),
            decode_urlsafe(salt_text),
            int(iterations_text),
        )
        return hmac.compare_digest(actual, decode_urlsafe(expected_text))
    except (KeyError, ValueError):
        return False


def create_session_token() -> str:
    expires_at = int(time.time()) + SESSION_TTL_SECONDS
    payload = f"{expires_at}:{secrets.token_urlsafe(12)}".encode("ascii")
    encoded_payload = base64.urlsafe_b64encode(payload).rstrip(b"=")
    signature = hmac.digest(os.environ["BANANA_WEB_SESSION_SECRET"].encode("utf-8"), encoded_payload, "sha256")
    encoded_signature = base64.urlsafe_b64encode(signature).rstrip(b"=")
    return f"{encoded_payload.decode('ascii')}.{encoded_signature.decode('ascii')}"


def session_token_is_valid(token: str) -> bool:
    try:
        encoded_payload, encoded_signature = token.split(".", 1)
        payload_bytes = encoded_payload.encode("ascii")
        expected = hmac.digest(os.environ["BANANA_WEB_SESSION_SECRET"].encode("utf-8"), payload_bytes, "sha256")
        if not hmac.compare_digest(expected, decode_urlsafe(encoded_signature)):
            return False
        expires_text, _ = decode_urlsafe(encoded_payload).decode("ascii").split(":", 1)
        return int(expires_text) >= int(time.time())
    except (KeyError, ValueError, UnicodeError):
        return False


def load_api_key(provider: dict[str, Any]) -> str:
    return get_api_key(provider["env"])


def public_job(job: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": job["id"],
        "provider": job["provider"],
        "provider_label": PROVIDERS[job["provider"]]["label"],
        "status": job["status"],
        "created_at": job["created_at"],
        "started_at": job.get("started_at"),
        "finished_at": job.get("finished_at"),
        "folder": job["folder"],
        "total": len(job["tasks"]),
        "completed": sum(1 for item in job["results"] if item["status"] in {"success", "failed"}),
        "results": [
            {
                key: value
                for key, value in result.items()
                if key not in {"image_bytes"}
            }
            for result in job["results"]
        ],
    }


def content_to_text(content: Any) -> str:
    if isinstance(content, str):
        return content
    if not isinstance(content, list):
        return ""
    parts: list[str] = []
    for part in content:
        if not isinstance(part, dict):
            continue
        if isinstance(part.get("text"), str):
            parts.append(part["text"])
        image_url = part.get("image_url")
        if isinstance(image_url, dict) and isinstance(image_url.get("url"), str):
            parts.append(image_url["url"])
    return "\n".join(parts)


def extract_image(content: str) -> tuple[bytes, str, str]:
    match = re.search(r"data:image/([^;]+);base64,([A-Za-z0-9+/=\r\n]+)", content)
    if not match:
        raise ValueError("模型响应中没有找到图片数据")
    raw_format = match.group(1).lower().strip()
    image_format = {"jpg": "jpeg", "jpeg": "jpeg", "png": "png", "webp": "webp", "gif": "gif"}.get(raw_format, "png")
    image_bytes = base64.b64decode(re.sub(r"\s+", "", match.group(2)), validate=True)
    if len(image_bytes) < 100:
        raise ValueError("模型返回的图片数据无效")
    text_only = re.sub(
        r"!?\[[^\]]*\]\(data:image/[^;]+;base64,[A-Za-z0-9+/=\r\n]+\)",
        "",
        content,
    ).strip()
    if text_only == content.strip():
        text_only = re.sub(r"data:image/[^;]+;base64,[A-Za-z0-9+/=\r\n]+", "", content).strip()
    return image_bytes, image_format, text_only


def decode_reference(image: dict[str, Any]) -> tuple[str, str, str]:
    name = str(image.get("name", "reference.png")).strip() or "reference.png"
    mime = str(image.get("mime", "image/png"))
    data = str(image.get("data", ""))
    if data.startswith("data:"):
        _, data = data.split(",", 1)
    decoded = base64.b64decode(data, validate=True)
    if not decoded:
        raise ValueError(f"参考图 {name} 为空")
    return name, mime, base64.b64encode(decoded).decode("ascii")


def generate_once(provider_id: str, prompt: str, images: list[dict[str, Any]]) -> tuple[bytes, str, str]:
    provider = PROVIDERS[provider_id]
    api_key = load_api_key(provider)
    if not api_key:
        raise ValueError(f"{provider['label']} 未配置 API Key")

    if provider.get("api_style") == "images":
        headers = {"Authorization": f"Bearer {api_key}", "Accept": "application/json"}
        if images:
            files = []
            for image in images:
                name, mime, encoded = decode_reference(image)
                files.append(("image[]", (name, base64.b64decode(encoded), mime)))
            response = requests.post(
                f"{provider['api_url'].rstrip('/')}/v1/images/edits",
                headers=headers,
                data={
                    "model": provider["model"],
                    "prompt": prompt,
                    "n": "1",
                    "size": "1024x1024",
                    "response_format": "b64_json",
                },
                files=files,
                timeout=600,
            )
        else:
            response = requests.post(
                f"{provider['api_url'].rstrip('/')}/v1/images/generations",
                headers={**headers, "Content-Type": "application/json; charset=utf-8"},
                json={
                    "model": provider["model"],
                    "prompt": prompt,
                    "n": 1,
                    "size": "1024x1024",
                    "response_format": "b64_json",
                },
                timeout=600,
            )
    else:
        content: list[dict[str, Any]] = [{"type": "text", "text": prompt}]
        for image in images:
            name, mime, encoded = decode_reference(image)
            content.append({"type": "text", "text": f"参考图文件名：{name}"})
            content.append({"type": "image_url", "image_url": {"url": f"data:{mime};base64,{encoded}"}})
        response = requests.post(
            provider["api_url"],
            headers={"Content-Type": "application/json", "Authorization": f"Bearer {api_key}"},
            json={
                "model": provider["model"],
                "stream": False,
                "messages": [{"role": "user", "content": content}],
            },
            timeout=300,
        )
    if response.status_code != 200:
        try:
            detail = response.json()
        except ValueError:
            detail = response.text[:500]
        raise RuntimeError(f"API 请求失败（{response.status_code}）：{detail}")
    try:
        body = response.json()
    except ValueError as exc:
        raise RuntimeError("API 返回了无法解析的数据") from exc

    if provider.get("api_style") == "images":
        data = body.get("data") or []
        item = data[0] if data and isinstance(data[0], dict) else {}
        encoded = item.get("b64_json")
        url = item.get("url")
        if isinstance(encoded, str):
            image_bytes = base64.b64decode(encoded, validate=True)
        elif isinstance(url, str) and url.startswith("data:image/") and "," in url:
            image_bytes = base64.b64decode(url.split(",", 1)[1])
        elif isinstance(url, str) and url.startswith(("https://", "http://")):
            image_response = requests.get(url, timeout=120)
            image_response.raise_for_status()
            image_bytes = image_response.content
        else:
            raise RuntimeError("API 响应中没有图片数据")
        if len(image_bytes) < 100:
            raise RuntimeError("API 返回的图片数据无效")
        if image_bytes.startswith(b"\xff\xd8\xff"):
            image_format = "jpeg"
        elif image_bytes.startswith(b"RIFF") and image_bytes[8:12] == b"WEBP":
            image_format = "webp"
        else:
            image_format = "png"
        return image_bytes, image_format, str(item.get("revised_prompt") or "")

    choices = body.get("choices") or []
    if not choices:
        raise RuntimeError("API 响应中没有 choices")
    message = choices[0].get("message") or {}
    content_text = content_to_text(message.get("content"))
    if not content_text:
        raise RuntimeError("模型没有返回内容")
    return extract_image(content_text)


def run_task(job: dict[str, Any], index: int, task: dict[str, Any]) -> None:
    with JOBS_LOCK:
        result = job["results"][index]
        if job["cancel_requested"]:
            result.update({"status": "stopped", "finished_at": now_iso()})
            return
        result["status"] = "running"
        result["started_at"] = now_iso()

    last_error = ""
    for attempt in range(1, 6):
        while not API_REQUEST_SLOTS.acquire(timeout=0.5):
            with JOBS_LOCK:
                if job["cancel_requested"]:
                    result.update({"status": "stopped", "finished_at": now_iso()})
                    return
        try:
            with JOBS_LOCK:
                if job["cancel_requested"]:
                    result.update({"status": "stopped", "finished_at": now_iso()})
                    return
                result["attempt"] = attempt
            image_bytes, image_format, response_text = generate_once(
                job["provider"], task["prompt"], task.get("images") or []
            )
            with JOBS_LOCK:
                result.update(
                    {
                        "status": "success",
                        "message": "生成成功",
                        "format": image_format,
                        "text": response_text,
                        "image_bytes": image_bytes,
                        "finished_at": now_iso(),
                    }
                )
            return
        except Exception as exc:  # Keep one failed task from stopping the batch.
            last_error = str(exc)
            with JOBS_LOCK:
                result["message"] = last_error
                cancelled = job["cancel_requested"]
            if cancelled:
                with JOBS_LOCK:
                    result.update({"status": "stopped", "finished_at": now_iso()})
                return
            if attempt < 5:
                time.sleep(5)
        finally:
            API_REQUEST_SLOTS.release()

    with JOBS_LOCK:
        result.update({"status": "failed", "message": last_error, "finished_at": now_iso()})


def run_job(job_id: str) -> None:
    with JOBS_LOCK:
        job = JOBS[job_id]
        job["status"] = "running"
        job["started_at"] = now_iso()

    worker_count = min(MAX_WORKERS_PER_JOB, len(job["tasks"]))
    with ThreadPoolExecutor(max_workers=worker_count, thread_name_prefix=f"job-{job_id[:8]}") as executor:
        futures = {
            executor.submit(run_task, job, index, task): index
            for index, task in enumerate(job["tasks"])
        }
        for future in as_completed(futures):
            try:
                future.result()
            except Exception as exc:
                with JOBS_LOCK:
                    result = job["results"][futures[future]]
                    result.update({"status": "failed", "message": str(exc), "finished_at": now_iso()})

    with JOBS_LOCK:
        stopped = any(result["status"] == "stopped" for result in job["results"])
        job["status"] = "stopped" if stopped else "completed"
        job["finished_at"] = now_iso()
        job["_finished_monotonic"] = time.monotonic()


def cleanup_jobs() -> None:
    cutoff = time.monotonic() - JOB_RETENTION_SECONDS
    with JOBS_LOCK:
        expired_ids = [
            job_id
            for job_id, job in JOBS.items()
            if job.get("_finished_monotonic", float("inf")) <= cutoff
        ]
        for job_id in expired_ids:
            del JOBS[job_id]


def cleanup_jobs_forever() -> None:
    while True:
        time.sleep(JOB_CLEANUP_INTERVAL_SECONDS)
        cleanup_jobs()


def parse_excel(encoded: str) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("缺少 openpyxl，请运行 py -m pip install openpyxl（或 python -m pip install openpyxl）") from exc
    try:
        raw = base64.b64decode(encoded, validate=True)
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"无法读取 Excel 文件：{exc}") from exc
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip().lower() for value in rows[0]]
    if "prompt" not in headers:
        raise ValueError("Excel 第一行必须包含 prompt 列")
    prompt_col = headers.index("prompt")
    image_col = headers.index("image_path") if "image_path" in headers else None
    tasks: list[dict[str, Any]] = []
    for row in rows[1:]:
        prompt = str(row[prompt_col] or "").strip() if prompt_col < len(row) else ""
        if not prompt:
            continue
        paths: list[str] = []
        if image_col is not None and image_col < len(row) and row[image_col] is not None:
            cell = str(row[image_col]).strip()
            if cell.startswith("["):
                try:
                    parsed = json.loads(cell)
                    paths = [str(item).strip() for item in parsed if str(item).strip()]
                except (ValueError, TypeError):
                    paths = []
            else:
                paths = [item.strip() for item in re.split(r"[;；]", cell) if item.strip()]
        tasks.append({"prompt": prompt, "image_path": paths})
    return tasks


class Handler(BaseHTTPRequestHandler):
    server_version = "BananaWeb/1.0"

    def log_message(self, fmt: str, *args: Any) -> None:
        print(f"[{self.log_date_time_string()}] {fmt % args}")

    def send_json(self, payload: Any, status: int = 200, headers: dict[str, str] | None = None) -> None:
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Cache-Control", "no-store")
        for name, value in (headers or {}).items():
            self.send_header(name, value)
        self.end_headers()
        self.wfile.write(data)

    def send_error_json(self, message: str, status: int = 400) -> None:
        self.send_json({"error": message}, status)

    def read_json(self) -> Any:
        length = int(self.headers.get("Content-Length", "0"))
        if length <= 0 or length > MAX_BODY_BYTES:
            raise ValueError("请求内容为空或过大")
        return json.loads(self.rfile.read(length).decode("utf-8"))

    def is_authenticated(self) -> bool:
        if not auth_enabled():
            return True
        cookie = SimpleCookie(self.headers.get("Cookie", ""))
        session = cookie.get(SESSION_COOKIE_NAME)
        return bool(session and session_token_is_valid(session.value))

    def redirect(self, location: str) -> None:
        self.send_response(303)
        self.send_header("Location", location)
        self.send_header("Cache-Control", "no-store")
        self.send_header("Content-Length", "0")
        self.end_headers()

    def require_authentication(self, path: str) -> bool:
        if self.is_authenticated():
            return True
        if path.startswith("/api/"):
            self.send_error_json("登录已失效，请重新登录", 401)
        else:
            self.redirect("/login")
        return False

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path in {"/login", "/login.html"}:
            if self.is_authenticated():
                self.redirect("/")
            else:
                self.serve_static("/login.html")
            return
        if parsed.path in {"/login.css", "/login.js"}:
            self.serve_static(parsed.path)
            return
        if not self.require_authentication(parsed.path):
            return
        if parsed.path == "/api/config":
            providers = [
                {
                    "id": key,
                    "label": value["label"],
                    "model": value["model"],
                    "configured": bool(load_api_key(value)),
                }
                for key, value in PROVIDERS.items()
            ]
            self.send_json({"providers": providers, "authentication_required": auth_enabled()})
            return

        job_match = re.fullmatch(r"/api/jobs/([a-f0-9]+)", parsed.path)
        if job_match:
            with JOBS_LOCK:
                job = JOBS.get(job_match.group(1))
                payload = public_job(job) if job else None
            if payload is None:
                self.send_error_json("任务不存在", 404)
            else:
                self.send_json(payload)
            return

        image_match = re.fullmatch(r"/api/jobs/([a-f0-9]+)/results/(\d+)/image", parsed.path)
        if image_match:
            with JOBS_LOCK:
                job = JOBS.get(image_match.group(1))
                index = int(image_match.group(2)) - 1
                result = job["results"][index] if job and 0 <= index < len(job["results"]) else None
                image_bytes = result.get("image_bytes") if result else None
                image_format = result.get("format", "png") if result else "png"
            if not image_bytes:
                self.send_error_json("图片尚未生成", 404)
                return
            self.send_response(200)
            self.send_header("Content-Type", f"image/{image_format}")
            self.send_header("Content-Length", str(len(image_bytes)))
            self.send_header("Cache-Control", "no-store")
            self.end_headers()
            self.wfile.write(image_bytes)
            return

        self.serve_static(parsed.path)

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        try:
            payload = self.read_json()
        except (ValueError, json.JSONDecodeError) as exc:
            self.send_error_json(str(exc))
            return

        if parsed.path == "/api/login":
            password = str(payload.get("password", ""))
            if not auth_enabled() or not password_matches(password):
                self.send_error_json("密码错误", 401)
                return
            cookie = (
                f"{SESSION_COOKIE_NAME}={create_session_token()}; Path=/; Max-Age={SESSION_TTL_SECONDS}; "
                "HttpOnly; Secure; SameSite=Strict"
            )
            self.send_json({"ok": True}, headers={"Set-Cookie": cookie})
            return

        if parsed.path == "/api/logout":
            cookie = f"{SESSION_COOKIE_NAME}=; Path=/; Max-Age=0; HttpOnly; Secure; SameSite=Strict"
            self.send_json({"ok": True}, headers={"Set-Cookie": cookie})
            return

        if not self.require_authentication(parsed.path):
            return

        if parsed.path == "/api/import-excel":
            try:
                tasks = parse_excel(str(payload.get("data", "")))
                self.send_json({"tasks": tasks})
            except ValueError as exc:
                self.send_error_json(str(exc))
            return

        if parsed.path == "/api/jobs":
            provider_id = str(payload.get("provider", ""))
            tasks = payload.get("tasks")
            if provider_id not in PROVIDERS:
                self.send_error_json("请选择有效的服务商")
                return
            if not load_api_key(PROVIDERS[provider_id]):
                self.send_error_json("所选服务商未配置 API Key")
                return
            if not isinstance(tasks, list) or not tasks:
                self.send_error_json("至少需要一条任务")
                return
            cleaned_tasks: list[dict[str, Any]] = []
            for task in tasks:
                prompt = str(task.get("prompt", "")).strip() if isinstance(task, dict) else ""
                images = task.get("images") if isinstance(task, dict) else []
                if not prompt:
                    self.send_error_json("所有任务都必须填写提示词")
                    return
                if not isinstance(images, list):
                    self.send_error_json("参考图数据格式错误")
                    return
                cleaned_tasks.append({"prompt": prompt, "images": images})
            job_id = uuid.uuid4().hex
            folder = dt.datetime.now().strftime("%m_%d_%H_%M_%S")
            job = {
                "id": job_id,
                "provider": provider_id,
                "status": "queued",
                "created_at": now_iso(),
                "folder": folder,
                "tasks": cleaned_tasks,
                "cancel_requested": False,
                "results": [
                    {
                        "index": index,
                        "prompt": task["prompt"],
                        "image_path": [image.get("relative_path", "") for image in task["images"]],
                        "status": "queued",
                        "attempt": 0,
                        "message": "等待中",
                    }
                    for index, task in enumerate(cleaned_tasks, 1)
                ],
            }
            with JOBS_LOCK:
                JOBS[job_id] = job
            threading.Thread(target=run_job, args=(job_id,), daemon=True).start()
            self.send_json({"job": public_job(job)}, 201)
            return

        stop_match = re.fullmatch(r"/api/jobs/([a-f0-9]+)/stop", parsed.path)
        if stop_match:
            with JOBS_LOCK:
                job = JOBS.get(stop_match.group(1))
                if job:
                    job["cancel_requested"] = True
            if not job:
                self.send_error_json("任务不存在", 404)
            else:
                self.send_json({"ok": True})
            return

        self.send_error_json("接口不存在", 404)

    def serve_static(self, request_path: str) -> None:
        relative = "index.html" if request_path in {"", "/"} else request_path.lstrip("/")
        candidate = (WEB_ROOT / relative).resolve()
        if WEB_ROOT.resolve() not in candidate.parents and candidate != WEB_ROOT.resolve():
            self.send_error(403)
            return
        if not candidate.is_file():
            self.send_error(404)
            return
        data = candidate.read_bytes()
        explicit_types = {
            ".html": "text/html",
            ".css": "text/css",
            ".js": "application/javascript",
        }
        content_type = explicit_types.get(candidate.suffix) or mimetypes.guess_type(candidate.name)[0] or "application/octet-stream"
        if candidate.suffix in {".html", ".css", ".js"}:
            content_type += "; charset=utf-8"
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Cache-Control", "no-cache")
        self.end_headers()
        self.wfile.write(data)


def existing_instance_url(port: int) -> str | None:
    url = f"http://127.0.0.1:{port}"
    try:
        response = requests.get(f"{url}/api/config", timeout=0.6)
        payload = response.json()
    except (requests.RequestException, ValueError):
        return None
    providers = payload.get("providers") if isinstance(payload, dict) else None
    if response.status_code == 200 and isinstance(providers, list):
        provider_ids = {item.get("id") for item in providers if isinstance(item, dict)}
        if {"banana", "ikun", "micu"}.issubset(provider_ids):
            return url
    return None


def main() -> None:
    parser = argparse.ArgumentParser(description="Banana local web UI")
    parser.add_argument("--port", type=int, default=8765)
    parser.add_argument("--no-browser", action="store_true")
    args = parser.parse_args()
    password_configured = bool(os.environ.get("BANANA_WEB_PASSWORD_HASH"))
    session_secret_configured = bool(os.environ.get("BANANA_WEB_SESSION_SECRET"))
    if password_configured != session_secret_configured:
        raise SystemExit("BANANA_WEB_PASSWORD_HASH 和 BANANA_WEB_SESSION_SECRET 必须同时配置")
    url = existing_instance_url(args.port)
    if url:
        print(f"Banana Web 已在运行：{url}")
        if not args.no_browser:
            webbrowser.open(url)
        return
    try:
        server = ThreadingHTTPServer(("127.0.0.1", args.port), Handler)
    except OSError as exc:
        raise SystemExit(f"本地端口 {args.port} 已被其他程序占用，请先关闭占用程序。") from exc
    url = f"http://127.0.0.1:{args.port}"
    print(f"Banana Web 已启动：{url}")
    print("关闭此窗口即可停止服务。")
    threading.Thread(target=cleanup_jobs_forever, daemon=True).start()
    if not args.no_browser:
        threading.Timer(0.6, lambda: webbrowser.open(url)).start()
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
